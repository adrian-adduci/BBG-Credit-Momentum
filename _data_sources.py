################################################################################
# Author: Adrian Adduci
# Email: FAA2160@columbia.edu
################################################################################
"""
Data source abstraction layer for loading economic data from multiple sources.

This module provides a unified interface for loading data from Excel files,
Bloomberg API, or other data sources. It allows easy switching between data
sources without changing the preprocessing pipeline.
"""
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Any
import logging
import pathlib

import pandas as pd
import numpy as np

logger = logging.getLogger("_data_sources")
logger.setLevel(logging.INFO)
path = pathlib.Path(__file__).parent.absolute()
handler = logging.FileHandler(path / "logs" / "_data_sources.log")
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)
logger.addHandler(handler)

################################################################################
# Security Definition Classes
################################################################################


@dataclass
class Security:
    """
    Universal security representation for any tradable instrument.

    This class provides a unified way to represent both traditional securities
    (Bloomberg credit indices, bonds) and crypto securities (BTC/USDT, ETH/USDT).

    Attributes:
        identifier: Security identifier (e.g., "BTC/USDT", "LF98TRUU Index")
        security_type: Type of security ("crypto_spot", "credit_index", "equity", "fx")
        source: Data source type ("binance", "bloomberg", "file")
        fields: List of data fields to fetch (e.g., ["close", "volume"] or ["OAS", "DTS"])
        metadata: Additional security-specific metadata
        features: Feature engineering configuration for this security

    Example:
        >>> # Crypto security
        >>> btc = Security(
        ...     identifier="BTC/USDT",
        ...     security_type="crypto_spot",
        ...     source="binance",
        ...     fields=["close", "volume"],
        ...     metadata={"exchange": "binance", "timeframe": "1h"}
        ... )
        >>>
        >>> # Credit index security
        >>> credit = Security(
        ...     identifier="LF98TRUU Index",
        ...     security_type="credit_index",
        ...     source="bloomberg",
        ...     fields=["OAS", "DTS"],
        ...     metadata={"display_name": "US Aggregate Bond Index"}
        ... )
    """

    identifier: str
    security_type: str
    source: str
    fields: List[str]
    metadata: Dict[str, Any] = field(default_factory=dict)
    features: Dict[str, Any] = field(default_factory=dict)

    def to_column_prefix(self) -> str:
        """
        Generate DataFrame column prefix from security identifier.

        Returns:
            str: Column prefix with special characters replaced

        Example:
            >>> Security("BTC/USDT", ...).to_column_prefix()
            "BTC_USDT"
            >>> Security("LF98TRUU Index", ...).to_column_prefix()
            "LF98TRUU_Index"
        """
        return self.identifier.replace("/", "_").replace(" ", "_").replace("-", "_")

    def get_column_names(self) -> List[str]:
        """
        Get list of DataFrame column names for this security.

        Returns:
            List[str]: Column names like ["BTC_USDT_close", "BTC_USDT_volume"]

        Example:
            >>> sec = Security("BTC/USDT", ..., fields=["close", "volume"])
            >>> sec.get_column_names()
            ["BTC_USDT_close", "BTC_USDT_volume"]
        """
        prefix = self.to_column_prefix()
        return [f"{prefix}_{field}" for field in self.fields]

    def is_crypto(self) -> bool:
        """Check if this is a cryptocurrency security."""
        return self.security_type.startswith("crypto")

    def is_credit(self) -> bool:
        """Check if this is a credit security."""
        return self.security_type in ["credit_index", "credit_bond"]

    def __repr__(self) -> str:
        return f"Security({self.identifier}, type={self.security_type}, source={self.source})"


################################################################################
# Abstract Base Class
################################################################################


class DataSource(ABC):
    """
    Abstract base class for all data sources.

    All data sources must implement the load_data() method which returns
    a pandas DataFrame with a standardized schema.

    Required columns in returned DataFrame:
        - Dates: datetime column with observation dates
        - Additional columns: numeric data for economic indicators
    """

    @abstractmethod
    def load_data(self) -> pd.DataFrame:
        """
        Load data from the source and return as a DataFrame.

        Returns:
            pd.DataFrame: Data with 'Dates' column and numeric indicator columns

        Raises:
            Exception: If data cannot be loaded
        """
        pass

    def validate_schema(self, df: pd.DataFrame, required_columns: Optional[List[str]] = None) -> None:
        """
        Validate that the DataFrame has the required schema.

        Args:
            df: DataFrame to validate
            required_columns: List of column names that must be present

        Raises:
            ValueError: If required columns are missing
        """
        if "Dates" not in df.columns:
            raise ValueError("DataFrame must contain a 'Dates' column")

        if required_columns:
            missing = set(required_columns) - set(df.columns)
            if missing:
                raise ValueError(f"Missing required columns: {missing}")

        logger.info(f"Schema validated successfully. Columns: {list(df.columns)}")


################################################################################
# Excel Data Source
################################################################################


class ExcelDataSource(DataSource):
    """
    Load data from Excel files (Bloomberg export format).

    Supports .xlsx and .xls files. Assumes first row contains column names
    and 'Dates' column contains dates in a parseable format.

    Args:
        file_path: Path to Excel file or file-like object
        sheet_name: Name or index of sheet to read (default: 0 = first sheet)
        date_column: Name of the date column (default: "Dates")

    Example:
        >>> source = ExcelDataSource("data/Economic_Data_2020_08_01.xlsx")
        >>> df = source.load_data()
        >>> print(df.head())
    """

    def __init__(
        self,
        file_path,
        sheet_name: int = 0,
        date_column: str = "Dates"
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.date_column = date_column
        logger.info(f"ExcelDataSource initialized with file: {file_path}")

    def load_data(self) -> pd.DataFrame:
        """
        Load data from Excel file.

        Returns:
            pd.DataFrame: Data with dates and economic indicators

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file cannot be read or schema is invalid
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            logger.info(f"Loaded {len(df)} rows from Excel file")

            # Ensure Dates column exists
            if self.date_column not in df.columns:
                raise ValueError(f"Date column '{self.date_column}' not found in Excel")

            # Standardize column name to 'Dates'
            if self.date_column != "Dates":
                df.rename(columns={self.date_column: "Dates"}, inplace=True)

            # Convert to datetime
            df["Dates"] = pd.to_datetime(df["Dates"])

            # Validate schema
            self.validate_schema(df)

            return df

        except FileNotFoundError as e:
            logger.error(f"Excel file not found: {self.file_path}")
            raise
        except Exception as e:
            logger.error(f"Failed to load Excel file: {str(e)}")
            raise ValueError(f"Failed to load Excel file: {str(e)}")


################################################################################
# Bloomberg API Exceptions
################################################################################


class BloombergAPIError(Exception):
    """Base exception for Bloomberg API errors."""
    pass


class BloombergTerminalNotRunning(BloombergAPIError):
    """Bloomberg Terminal is not running or not accessible."""
    pass


class BloombergAuthenticationError(BloombergAPIError):
    """Bloomberg authentication failed."""
    pass


class BloombergInvalidSecurity(BloombergAPIError):
    """Invalid Bloomberg security identifier."""
    pass


class BloombergFieldError(BloombergAPIError):
    """Bloomberg field not available for security."""
    pass


################################################################################
# Bloomberg API Data Source
################################################################################


class BloombergAPIDataSource(DataSource):
    """
    Load data from Bloomberg Terminal API (requires blpapi library).

    This implementation provides production-ready Bloomberg Terminal integration
    with comprehensive error handling, retry logic, and batch processing.

    Args:
        securities: List of Bloomberg security tickers (e.g., ["LF98TRUU Index"])
        fields: List of Bloomberg fields to retrieve (e.g., ["OAS", "PX_LAST"])
        start_date: Start date for historical data
        end_date: End date for historical data
        host: Bloomberg API host (default: "localhost")
        port: Bloomberg API port (default: 8194)
        timeout: Connection timeout in milliseconds (default: 30000)
        max_retries: Maximum number of retry attempts (default: 3)
        batch_size: Max securities per request (default: 100)

    Example:
        >>> source = BloombergAPIDataSource(
        ...     securities=["LF98TRUU Index", "LUACTRUU Index"],
        ...     fields=["OAS", "PX_LAST"],
        ...     start_date=datetime(2020, 1, 1),
        ...     end_date=datetime(2020, 12, 31)
        ... )
        >>> df = source.load_data()
    """

    SUPPORTED_FIELDS = {
        "OAS": "Option-Adjusted Spread",
        "PX_LAST": "Last Price",
        "PX_OPEN": "Open Price",
        "PX_HIGH": "High Price",
        "PX_LOW": "Low Price",
        "PX_VOLUME": "Volume",
        "YLD_YTM_MID": "Yield to Maturity",
        "DTS": "Duration to Worst",
        "AMOUNT_OUTSTANDING": "Outstanding Amount",
        "RTG_MOODY": "Moody's Rating",
        "RTG_SP": "S&P Rating",
    }

    def __init__(
        self,
        securities: List[str],
        fields: List[str],
        start_date: datetime,
        end_date: datetime,
        host: str = "localhost",
        port: int = 8194,
        timeout: int = 30000,
        max_retries: int = 3,
        batch_size: int = 100
    ):
        self.securities = securities
        self.fields = fields
        self.start_date = start_date
        self.end_date = end_date
        self.host = host
        self.port = port
        self.timeout = timeout
        self.max_retries = max_retries
        self.batch_size = batch_size
        self.session = None
        logger.info(
            f"BloombergAPIDataSource initialized: {len(securities)} securities, "
            f"{len(fields)} fields, {start_date.date()} to {end_date.date()}"
        )

    def _import_blpapi(self):
        """Import Bloomberg API library with helpful error message."""
        try:
            import blpapi
            return blpapi
        except ImportError:
            error_msg = (
                "Bloomberg API (blpapi) not installed. "
                "Install with: pip install --index-url=https://bcms.bloomberg.com/pip/simple blpapi\n"
                "Note: Requires Bloomberg Terminal installation and DAPI<GO> setup."
            )
            logger.error(error_msg)
            raise ImportError(error_msg)

    def _create_session(self):
        """
        Create and start Bloomberg API session.

        Returns:
            blpapi.Session: Active Bloomberg session

        Raises:
            BloombergTerminalNotRunning: If Terminal is not accessible
            BloombergAuthenticationError: If authentication fails
        """
        blpapi = self._import_blpapi()

        logger.info(f"Connecting to Bloomberg Terminal at {self.host}:{self.port}")

        # Configure session options
        session_options = blpapi.SessionOptions()
        session_options.setServerHost(self.host)
        session_options.setServerPort(self.port)

        # Create session
        session = blpapi.Session(session_options)

        # Start session with timeout
        if not session.start():
            error_msg = (
                f"Failed to start Bloomberg session at {self.host}:{self.port}. "
                "Ensure Bloomberg Terminal is running and DAPI<GO> is configured."
            )
            logger.error(error_msg)
            raise BloombergTerminalNotRunning(error_msg)

        # Open reference data service
        if not session.openService("//blp/refdata"):
            session.stop()
            error_msg = "Failed to open Bloomberg reference data service"
            logger.error(error_msg)
            raise BloombergAuthenticationError(error_msg)

        logger.info("Successfully connected to Bloomberg Terminal")
        return session

    def _create_historical_request(self, service, securities: List[str]):
        """
        Create HistoricalDataRequest for given securities.

        Args:
            service: Bloomberg refdata service
            securities: List of security identifiers

        Returns:
            blpapi.Request: Configured historical data request
        """
        request = service.createRequest("HistoricalDataRequest")

        # Add securities
        for security in securities:
            request.append("securities", security)

        # Add fields
        for field in self.fields:
            request.append("fields", field)

        # Set date range
        request.set("startDate", self.start_date.strftime("%Y%m%d"))
        request.set("endDate", self.end_date.strftime("%Y%m%d"))

        # Optional: Add additional settings
        request.set("periodicitySelection", "DAILY")

        logger.debug(f"Created request for {len(securities)} securities: {securities}")
        return request

    def _parse_response(self, response_event, blpapi) -> Dict[str, Dict[str, List]]:
        """
        Parse Bloomberg API response into structured data.

        Args:
            response_event: Bloomberg response event
            blpapi: Bloomberg API module

        Returns:
            Dict mapping security -> field -> [values]
        """
        data = {}

        for msg in response_event:
            if msg.hasElement("responseError"):
                error = msg.getElement("responseError")
                logger.warning(f"Response error: {error}")
                continue

            security_data = msg.getElement("securityData")
            security = security_data.getElementAsString("security")

            # Check for security errors
            if security_data.hasElement("securityError"):
                error = security_data.getElement("securityError")
                error_msg = f"Security error for {security}: {error}"
                logger.warning(error_msg)
                continue

            # Initialize data structure for this security
            if security not in data:
                data[security] = {"date": []}
                for field in self.fields:
                    data[security][field] = []

            # Parse field data
            field_data_array = security_data.getElement("fieldData")
            for i in range(field_data_array.numValues()):
                field_data = field_data_array.getValueAsElement(i)

                # Get date
                date = field_data.getElementAsDatetime("date")
                data[security]["date"].append(pd.Timestamp(date))

                # Get field values
                for field in self.fields:
                    if field_data.hasElement(field):
                        value = field_data.getElementAsFloat(field)
                        data[security][field].append(value)
                    else:
                        # Field not available for this date
                        data[security][field].append(None)

        return data

    def _fetch_data_batch(self, session, service, securities: List[str], retry_count: int = 0):
        """
        Fetch data for a batch of securities with retry logic.

        Args:
            session: Active Bloomberg session
            service: Bloomberg refdata service
            securities: List of securities for this batch
            retry_count: Current retry attempt

        Returns:
            Dict of parsed data for this batch
        """
        blpapi = self._import_blpapi()

        try:
            # Create and send request
            request = self._create_historical_request(service, securities)
            session.sendRequest(request)

            # Process responses
            batch_data = {}
            while True:
                event = session.nextEvent(self.timeout)

                if event.eventType() == blpapi.Event.RESPONSE:
                    # Final response
                    parsed = self._parse_response(event, blpapi)
                    batch_data.update(parsed)
                    break
                elif event.eventType() == blpapi.Event.PARTIAL_RESPONSE:
                    # Intermediate response
                    parsed = self._parse_response(event, blpapi)
                    batch_data.update(parsed)
                elif event.eventType() == blpapi.Event.TIMEOUT:
                    raise TimeoutError(f"Request timeout after {self.timeout}ms")
                else:
                    # Other event types (REQUEST_STATUS, etc.)
                    continue

            return batch_data

        except Exception as e:
            if retry_count < self.max_retries:
                logger.warning(
                    f"Batch request failed (attempt {retry_count + 1}/{self.max_retries}): {e}"
                )
                import time
                time.sleep(2 ** retry_count)  # Exponential backoff
                return self._fetch_data_batch(session, service, securities, retry_count + 1)
            else:
                logger.error(f"Batch request failed after {self.max_retries} retries: {e}")
                raise

    def load_data(self) -> pd.DataFrame:
        """
        Load data from Bloomberg Terminal API.

        Returns:
            pd.DataFrame: Data with 'Dates' column and field columns per security

        Raises:
            ImportError: If blpapi is not installed
            BloombergTerminalNotRunning: If Terminal not accessible
            BloombergAPIError: If data cannot be retrieved
        """
        all_data = {}

        try:
            # Create session
            self.session = self._create_session()
            service = self.session.getService("//blp/refdata")

            # Process securities in batches
            for i in range(0, len(self.securities), self.batch_size):
                batch = self.securities[i:i + self.batch_size]
                logger.info(
                    f"Fetching batch {i // self.batch_size + 1} "
                    f"({len(batch)} securities)"
                )

                batch_data = self._fetch_data_batch(self.session, service, batch)
                all_data.update(batch_data)

            # Convert to DataFrame
            df = self._build_dataframe(all_data)

            # Validate
            self.validate_schema(df)

            logger.info(
                f"Successfully loaded {len(df)} rows, {len(df.columns) - 1} columns "
                f"from Bloomberg API"
            )
            return df

        finally:
            # Always close session
            if self.session:
                self.session.stop()
                logger.info("Bloomberg session closed")

    def _build_dataframe(self, data: Dict[str, Dict[str, List]]) -> pd.DataFrame:
        """
        Build unified DataFrame from parsed Bloomberg data.

        Args:
            data: Dict mapping security -> field -> [values]

        Returns:
            pd.DataFrame with 'Dates' column and {security}_{field} columns
        """
        if not data:
            raise ValueError("No data retrieved from Bloomberg API")

        # Collect all unique dates across all securities
        all_dates = set()
        for security_data in data.values():
            all_dates.update(security_data["date"])

        # Create DataFrame with dates
        df = pd.DataFrame({"Dates": sorted(all_dates)})

        # Add columns for each security-field combination
        for security, security_data in data.items():
            # Create temporary DataFrame for this security
            temp_df = pd.DataFrame({
                "Dates": security_data["date"]
            })

            for field in self.fields:
                col_name = f"{security.replace(' ', '_')}_{field}"
                temp_df[col_name] = security_data[field]

            # Merge with main DataFrame
            df = df.merge(temp_df, on="Dates", how="left")

        # Sort by date
        df = df.sort_values("Dates").reset_index(drop=True)

        logger.debug(f"Built DataFrame with shape {df.shape}")
        return df


################################################################################
# CSV Data Source
################################################################################


class CSVDataSource(DataSource):
    """
    Load data from CSV files.

    Args:
        file_path: Path to CSV file
        date_column: Name of the date column (default: "Dates")
        date_format: Datetime format string (default: None = auto-detect)

    Example:
        >>> source = CSVDataSource("data/economic_data.csv")
        >>> df = source.load_data()
    """

    def __init__(
        self,
        file_path: str,
        date_column: str = "Dates",
        date_format: Optional[str] = None
    ):
        self.file_path = file_path
        self.date_column = date_column
        self.date_format = date_format
        logger.info(f"CSVDataSource initialized with file: {file_path}")

    def load_data(self) -> pd.DataFrame:
        """
        Load data from CSV file.

        Returns:
            pd.DataFrame: Data with dates and indicators

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If schema is invalid
        """
        try:
            df = pd.read_csv(self.file_path)
            logger.info(f"Loaded {len(df)} rows from CSV file")

            # Ensure Dates column exists
            if self.date_column not in df.columns:
                raise ValueError(f"Date column '{self.date_column}' not found in CSV")

            # Standardize column name
            if self.date_column != "Dates":
                df.rename(columns={self.date_column: "Dates"}, inplace=True)

            # Convert to datetime
            if self.date_format:
                df["Dates"] = pd.to_datetime(df["Dates"], format=self.date_format)
            else:
                df["Dates"] = pd.to_datetime(df["Dates"])

            self.validate_schema(df)

            return df

        except FileNotFoundError as e:
            logger.error(f"CSV file not found: {self.file_path}")
            raise
        except Exception as e:
            logger.error(f"Failed to load CSV file: {str(e)}")
            raise ValueError(f"Failed to load CSV file: {str(e)}")


################################################################################
# Bloomberg Excel Data Source (Enhanced)
################################################################################


class BloombergExcelDataSource(ExcelDataSource):
    """
    Enhanced Excel data source with Bloomberg-specific schema validation.

    Handles Bloomberg Excel exports with:
    - Bloomberg error value detection (#N/A N/A, #VALUE!)
    - Formula detection (=BDH, =BDP)
    - Multi-sheet support (timeseries + metadata)
    - Schema validation for Bloomberg format

    Args:
        file_path: Path to Excel file with Bloomberg data
        sheet_name: Sheet name or index (default: 0)
        date_column: Name of date column (default: "Dates")
        handle_errors: Convert Bloomberg errors to NaN (default: True)
        validate_formulas: Check for Bloomberg formulas (default: False)

    Example:
        >>> source = BloombergExcelDataSource(
        ...     file_path="data/bloomberg_export.xlsx",
        ...     handle_errors=True
        ... )
        >>> df = source.load_data()
    """

    BLOOMBERG_ERROR_VALUES = [
        "#N/A N/A",
        "#N/A Field Not Applicable",
        "#N/A Invalid Security",
        "#VALUE!",
        "#REF!",
        "#NAME?",
        "#NUM!",
        "#DIV/0!",
    ]

    def __init__(
        self,
        file_path,
        sheet_name: int = 0,
        date_column: str = "Dates",
        handle_errors: bool = True,
        validate_formulas: bool = False
    ):
        super().__init__(file_path, sheet_name, date_column)
        self.handle_errors = handle_errors
        self.validate_formulas = validate_formulas

    def load_data(self) -> pd.DataFrame:
        """
        Load Bloomberg Excel data with error handling.

        Returns:
            pd.DataFrame: Cleaned data with Bloomberg errors converted to NaN

        Raises:
            ValueError: If schema validation fails
        """
        try:
            # Load Excel file
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            logger.info(f"Loaded {len(df)} rows from Bloomberg Excel file")

            # Handle Bloomberg error values
            if self.handle_errors:
                df = self._handle_bloomberg_errors(df)

            # Ensure Dates column exists
            if self.date_column not in df.columns:
                raise ValueError(f"Date column '{self.date_column}' not found in Excel")

            # Standardize column name
            if self.date_column != "Dates":
                df.rename(columns={self.date_column: "Dates"}, inplace=True)

            # Convert to datetime
            df["Dates"] = pd.to_datetime(df["Dates"])

            # Validate Bloomberg schema
            self._validate_bloomberg_schema(df)

            # Validate schema
            self.validate_schema(df)

            return df

        except FileNotFoundError as e:
            logger.error(f"Bloomberg Excel file not found: {self.file_path}")
            raise
        except Exception as e:
            logger.error(f"Failed to load Bloomberg Excel file: {str(e)}")
            raise ValueError(f"Failed to load Bloomberg Excel file: {str(e)}")

    def _handle_bloomberg_errors(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Replace Bloomberg error values with NaN.

        Args:
            df: DataFrame with potential Bloomberg errors

        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        error_count = 0
        for col in df.columns:
            if col == self.date_column:
                continue

            # Replace error strings with NaN
            for error_val in self.BLOOMBERG_ERROR_VALUES:
                mask = df[col].astype(str).str.contains(error_val, na=False)
                error_count += mask.sum()
                df.loc[mask, col] = pd.NA

        if error_count > 0:
            logger.warning(
                f"Converted {error_count} Bloomberg error values to NaN"
            )

        return df

    def _validate_bloomberg_schema(self, df: pd.DataFrame) -> None:
        """
        Validate that DataFrame follows Bloomberg export format.

        Args:
            df: DataFrame to validate

        Raises:
            ValueError: If schema validation fails
        """
        issues = []

        # Check for dates in ascending order
        if not df["Dates"].is_monotonic_increasing:
            issues.append("Dates are not in ascending order")

        # Check for duplicate dates
        if df["Dates"].duplicated().any():
            dup_count = df["Dates"].duplicated().sum()
            issues.append(f"Found {dup_count} duplicate dates")

        # Check for at least one data column
        data_cols = [c for c in df.columns if c != "Dates"]
        if len(data_cols) == 0:
            issues.append("No data columns found (only Dates)")

        # Warn about columns with all NaN
        for col in data_cols:
            if df[col].isna().all():
                logger.warning(f"Column '{col}' contains all NaN values")

        if issues:
            error_msg = "Bloomberg schema validation failed:\n" + "\n".join(f"  - {issue}" for issue in issues)
            logger.error(error_msg)
            raise ValueError(error_msg)

        logger.info("Bloomberg Excel schema validated successfully")

    def detect_bloomberg_formulas(self) -> Dict[str, str]:
        """
        Detect Bloomberg formulas in Excel file (requires openpyxl).

        Returns:
            Dict mapping column names to detected formulas

        Note:
            This requires reading the Excel file with formulas preserved.
            Only works with .xlsx files.
        """
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(self.file_path)
            sheet = workbook.active if self.sheet_name == 0 else workbook[self.sheet_name]

            formulas = {}
            # Check first data row (row 2, assuming row 1 is headers)
            for cell in sheet[2]:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith("=BDH") or cell.value.startswith("=BDP"):
                        col_name = sheet.cell(1, cell.column).value
                        formulas[col_name] = cell.value

            if formulas:
                logger.info(f"Detected {len(formulas)} Bloomberg formulas")

            return formulas

        except ImportError:
            logger.warning("openpyxl not available, cannot detect formulas")
            return {}
        except Exception as e:
            logger.warning(f"Failed to detect formulas: {e}")
            return {}


################################################################################
# Hybrid Bloomberg Data Source (API + Excel Fallback)
################################################################################


class HybridBloombergDataSource(DataSource):
    """
    Intelligent Bloomberg data source with API and Excel fallback.

    Attempts to load data from Bloomberg API first. If API is unavailable
    (Terminal not running, connection issues), falls back to Excel file.

    This provides maximum reliability for production systems where
    Bloomberg Terminal may not always be available.

    Args:
        securities: List of Bloomberg security tickers
        fields: List of Bloomberg fields
        start_date: Start date for historical data
        end_date: End date for historical data
        excel_fallback_path: Path to Excel file for fallback
        api_host: Bloomberg API host (default: "localhost")
        api_port: Bloomberg API port (default: 8194)
        prefer_api: Try API first if True, Excel first if False (default: True)

    Example:
        >>> source = HybridBloombergDataSource(
        ...     securities=["LF98TRUU Index"],
        ...     fields=["OAS"],
        ...     start_date=datetime(2020, 1, 1),
        ...     end_date=datetime(2020, 12, 31),
        ...     excel_fallback_path="data/bloomberg_export.xlsx",
        ... )
        >>> df = source.load_data()
    """

    def __init__(
        self,
        securities: List[str],
        fields: List[str],
        start_date: datetime,
        end_date: datetime,
        excel_fallback_path: str,
        api_host: str = "localhost",
        api_port: int = 8194,
        prefer_api: bool = True
    ):
        self.securities = securities
        self.fields = fields
        self.start_date = start_date
        self.end_date = end_date
        self.excel_fallback_path = excel_fallback_path
        self.api_host = api_host
        self.api_port = api_port
        self.prefer_api = prefer_api

        # Initialize data sources
        self.api_source = BloombergAPIDataSource(
            securities=securities,
            fields=fields,
            start_date=start_date,
            end_date=end_date,
            host=api_host,
            port=api_port
        )

        self.excel_source = BloombergExcelDataSource(
            file_path=excel_fallback_path
        )

        logger.info(
            f"HybridBloombergDataSource initialized with "
            f"{'API-first' if prefer_api else 'Excel-first'} strategy"
        )

    def load_data(self) -> pd.DataFrame:
        """
        Load data from Bloomberg, trying API then Excel (or vice versa).

        Returns:
            pd.DataFrame: Bloomberg data from either source

        Raises:
            Exception: If both API and Excel fail
        """
        if self.prefer_api:
            # Try API first
            try:
                logger.info("Attempting to load from Bloomberg API...")
                df = self.api_source.load_data()
                logger.info("Successfully loaded data from Bloomberg API")
                return df
            except (BloombergTerminalNotRunning, ImportError) as e:
                logger.warning(f"Bloomberg API unavailable: {e}")
                logger.info("Falling back to Excel file...")
                try:
                    df = self.excel_source.load_data()
                    logger.info("Successfully loaded data from Excel fallback")
                    return df
                except Exception as excel_error:
                    logger.error(f"Excel fallback also failed: {excel_error}")
                    raise Exception(
                        f"Both Bloomberg API and Excel fallback failed. "
                        f"API error: {e}, Excel error: {excel_error}"
                    )
            except Exception as e:
                logger.error(f"Bloomberg API error: {e}")
                logger.info("Falling back to Excel file...")
                try:
                    df = self.excel_source.load_data()
                    logger.info("Successfully loaded data from Excel fallback")
                    return df
                except Exception as excel_error:
                    logger.error(f"Excel fallback also failed: {excel_error}")
                    raise Exception(
                        f"Both Bloomberg API and Excel fallback failed. "
                        f"API error: {e}, Excel error: {excel_error}"
                    )
        else:
            # Try Excel first
            try:
                logger.info("Loading from Excel file...")
                df = self.excel_source.load_data()
                logger.info("Successfully loaded data from Excel")
                return df
            except Exception as excel_error:
                logger.warning(f"Excel loading failed: {excel_error}")
                logger.info("Falling back to Bloomberg API...")
                try:
                    df = self.api_source.load_data()
                    logger.info("Successfully loaded data from Bloomberg API")
                    return df
                except Exception as api_error:
                    logger.error(f"Bloomberg API also failed: {api_error}")
                    raise Exception(
                        f"Both Excel and Bloomberg API failed. "
                        f"Excel error: {excel_error}, API error: {api_error}"
                    )


################################################################################
# Mixed Portfolio Data Source (Unified Crypto + Traditional)
################################################################################


class MixedPortfolioDataSource(DataSource):
    """
    Unified data source for mixed portfolios (crypto + traditional securities).

    This class combines multiple data sources (Bloomberg, crypto exchanges,
    blockchain metrics) into a single DataFrame with aligned dates.

    Features:
    - Loads data from multiple sources in parallel
    - Aligns dates across different market calendars (24/7 crypto vs weekday credit)
    - Forward fills missing data with configurable limits
    - Validates data quality across all securities
    - Provides unified column naming

    Args:
        securities: List of Security objects to load
        start_date: Start date for data range
        end_date: End date for data range
        alignment_method: How to align dates ("outer", "inner", "left")
        fill_method: Method to fill missing values ("ffill", "bfill", None)
        fill_limit: Maximum number of periods to forward/backward fill
        validate: Whether to validate data after loading

    Example:
        >>> securities = [
        ...     Security("BTC/USDT", "crypto_spot", "binance", ["close"]),
        ...     Security("LF98TRUU Index", "credit_index", "bloomberg", ["OAS"])
        ... ]
        >>> source = MixedPortfolioDataSource(
        ...     securities=securities,
        ...     start_date=datetime(2020, 1, 1),
        ...     end_date=datetime(2024, 12, 31),
        ...     alignment_method="outer",
        ...     fill_method="ffill",
        ...     fill_limit=5
        ... )
        >>> df = source.load_data()
        >>> print(df.columns)
        ['Dates', 'BTC_USDT_close', 'LF98TRUU_Index_OAS']
    """

    def __init__(
        self,
        securities: List[Security],
        start_date: datetime,
        end_date: datetime,
        alignment_method: str = "outer",
        fill_method: Optional[str] = "ffill",
        fill_limit: int = 5,
        validate: bool = True
    ):
        self.securities = securities
        self.start_date = start_date
        self.end_date = end_date
        self.alignment_method = alignment_method
        self.fill_method = fill_method
        self.fill_limit = fill_limit
        self.validate = validate

        logger.info(
            f"MixedPortfolioDataSource initialized with {len(securities)} securities "
            f"({sum(1 for s in securities if s.is_crypto())} crypto, "
            f"{sum(1 for s in securities if s.is_credit())} credit)"
        )

    def load_data(self) -> pd.DataFrame:
        """
        Load and merge data from all securities.

        Returns:
            pd.DataFrame: Unified DataFrame with aligned dates and all securities

        Raises:
            ValueError: If no securities provided or all sources fail
        """
        if not self.securities:
            raise ValueError("No securities provided to load")

        # Load data from each security
        security_dataframes = {}
        for security in self.securities:
            try:
                logger.info(f"Loading data for {security.identifier}...")
                df = self._load_security_data(security)
                security_dataframes[security.identifier] = df
                logger.info(
                    f"Loaded {len(df)} rows for {security.identifier}"
                )
            except Exception as e:
                logger.error(
                    f"Failed to load {security.identifier}: {e}"
                )
                # Continue with other securities

        if not security_dataframes:
            raise ValueError("Failed to load data from all securities")

        # Merge all DataFrames
        logger.info("Merging data from all securities...")
        merged_df = self._merge_dataframes(security_dataframes)

        # Fill missing values
        if self.fill_method:
            logger.info(
                f"Filling missing values with {self.fill_method} "
                f"(limit={self.fill_limit})"
            )
            merged_df = self._fill_missing_values(merged_df)

        # Validate data
        if self.validate:
            self._validate_data(merged_df)

        logger.info(
            f"Successfully loaded mixed portfolio: {len(merged_df)} rows, "
            f"{len(merged_df.columns) - 1} columns"
        )

        return merged_df

    def _load_security_data(self, security: Security) -> pd.DataFrame:
        """
        Load data for a single security based on its source.

        Args:
            security: Security object to load

        Returns:
            pd.DataFrame: Data for this security with Dates column
        """
        source_type = security.source.lower()

        # Determine data source based on security source
        if source_type in ["binance", "coinbase", "kraken", "bybit", "okx"]:
            # Crypto exchange
            from _crypto_data_sources import CryptoExchangeDataSource
            source = CryptoExchangeDataSource(
                exchange_id=source_type,
                symbols=[security.identifier],
                timeframe=security.metadata.get("timeframe", "1h"),
                start_date=self.start_date,
                end_date=self.end_date
            )
            df = source.load_data()

        elif source_type == "bloomberg":
            # Bloomberg API or Excel
            excel_fallback = security.metadata.get("excel_fallback")
            if excel_fallback:
                # Use hybrid mode
                source = HybridBloombergDataSource(
                    securities=[security.identifier],
                    fields=security.fields,
                    start_date=self.start_date,
                    end_date=self.end_date,
                    excel_fallback_path=excel_fallback
                )
            else:
                # API only
                source = BloombergAPIDataSource(
                    securities=[security.identifier],
                    fields=security.fields,
                    start_date=self.start_date,
                    end_date=self.end_date
                )
            df = source.load_data()

        elif source_type == "bloomberg_excel":
            # Bloomberg Excel file
            source = BloombergExcelDataSource(
                file_path=security.metadata["file_path"]
            )
            df = source.load_data()

        elif source_type == "excel":
            # Generic Excel file
            source = ExcelDataSource(
                file_path=security.metadata["file_path"]
            )
            df = source.load_data()

        elif source_type == "csv":
            # CSV file
            source = CSVDataSource(
                file_path=security.metadata["file_path"]
            )
            df = source.load_data()

        elif source_type == "blockchain":
            # Blockchain on-chain metrics
            from data_sources.blockchain_provider import BlockchainDataSource
            provider = security.metadata.get("provider", "glassnode")
            asset = security.identifier.split("/")[0] if "/" in security.identifier else security.identifier
            source = BlockchainDataSource(
                provider=provider,
                assets=[asset],
                metrics=security.fields,
                start_date=self.start_date,
                end_date=self.end_date
            )
            df = source.load_data()

        else:
            raise ValueError(f"Unknown source type: {source_type}")

        return df

    def _merge_dataframes(
        self,
        security_dataframes: Dict[str, pd.DataFrame]
    ) -> pd.DataFrame:
        """
        Merge multiple security DataFrames on Dates column.

        Args:
            security_dataframes: Dict mapping security ID to its DataFrame

        Returns:
            pd.DataFrame: Merged DataFrame with all securities
        """
        # Start with dates from first security
        first_security = list(security_dataframes.values())[0]
        merged_df = first_security[["Dates"]].copy()

        # Merge each security's data
        for security_id, df in security_dataframes.items():
            # Merge on Dates
            merged_df = merged_df.merge(
                df,
                on="Dates",
                how=self.alignment_method,
                suffixes=("", f"_{security_id}")
            )

        # Sort by date
        merged_df = merged_df.sort_values("Dates").reset_index(drop=True)

        return merged_df

    def _fill_missing_values(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Fill missing values using specified method.

        Args:
            df: DataFrame with potential missing values

        Returns:
            pd.DataFrame: DataFrame with filled values
        """
        # Don't fill Dates column
        data_cols = [c for c in df.columns if c != "Dates"]

        if self.fill_method == "ffill":
            # Forward fill
            df[data_cols] = df[data_cols].fillna(method="ffill", limit=self.fill_limit)
        elif self.fill_method == "bfill":
            # Backward fill
            df[data_cols] = df[data_cols].fillna(method="bfill", limit=self.fill_limit)

        # Log remaining missing values
        missing_count = df[data_cols].isna().sum().sum()
        if missing_count > 0:
            logger.warning(
                f"{missing_count} missing values remain after filling "
                f"(fill_limit={self.fill_limit})"
            )

        return df

    def _validate_data(self, df: pd.DataFrame) -> None:
        """
        Validate merged data quality.

        Args:
            df: DataFrame to validate

        Raises:
            ValueError: If validation fails
        """
        issues = []

        # Check for dates
        if "Dates" not in df.columns:
            issues.append("Missing 'Dates' column")

        # Check for data columns
        data_cols = [c for c in df.columns if c != "Dates"]
        if len(data_cols) == 0:
            issues.append("No data columns found")

        # Check date order
        if not df["Dates"].is_monotonic_increasing:
            issues.append("Dates are not in ascending order")

        # Check for duplicate dates
        if df["Dates"].duplicated().any():
            dup_count = df["Dates"].duplicated().sum()
            issues.append(f"Found {dup_count} duplicate dates")

        # Warn about high missing value percentage
        for col in data_cols:
            missing_pct = df[col].isna().sum() / len(df) * 100
            if missing_pct > 50:
                logger.warning(
                    f"Column '{col}' has {missing_pct:.1f}% missing values"
                )

        if issues:
            error_msg = "Data validation failed:\n" + "\n".join(f"  - {issue}" for issue in issues)
            logger.error(error_msg)
            raise ValueError(error_msg)

        logger.info("Data validation passed")


################################################################################
# Data Source Factory
################################################################################


class DataSourceFactory:
    """
    Factory class for creating data source instances.

    Simplifies data source creation by providing a unified interface.

    Example:
        >>> # Create from Excel
        >>> source = DataSourceFactory.create("excel", file_path="data.xlsx")
        >>>
        >>> # Create from CSV
        >>> source = DataSourceFactory.create("csv", file_path="data.csv")
        >>>
        >>> # Create from Bloomberg API (when implemented)
        >>> source = DataSourceFactory.create(
        ...     "bloomberg",
        ...     securities=["LF98TRUU Index"],
        ...     fields=["OAS"],
        ...     start_date=datetime(2020, 1, 1),
        ...     end_date=datetime(2020, 12, 31)
        ... )
        >>>
        >>> # Create from Blockchain provider
        >>> source = DataSourceFactory.create(
        ...     "blockchain",
        ...     provider="glassnode",
        ...     assets=["BTC", "ETH"],
        ...     metrics=["mvrv", "nvt", "active_addresses"],
        ...     start_date=datetime(2024, 1, 1),
        ...     end_date=datetime.now()
        ... )
    """

    @staticmethod
    def create(source_type: str, **kwargs) -> DataSource:
        """
        Create a data source instance.

        Args:
            source_type: Type of data source
            **kwargs: Arguments to pass to the data source constructor

        Returns:
            DataSource: Instance of the requested data source type

        Raises:
            ValueError: If source_type is not recognized

        Supported source types:
            - excel: Generic Excel file
            - csv: CSV file
            - bloomberg: Bloomberg Terminal API
            - bloomberg_excel: Bloomberg Excel export (enhanced)
            - bloomberg_hybrid: Bloomberg with API/Excel fallback
            - crypto: Cryptocurrency exchange data
            - crypto_ws: Crypto WebSocket (real-time)
            - crypto_agg: Multi-exchange aggregator
            - blockchain: Blockchain on-chain metrics
            - mixed_portfolio: Mixed crypto + traditional securities
        """
        source_type = source_type.lower()

        if source_type == "excel":
            return ExcelDataSource(**kwargs)
        elif source_type == "csv":
            return CSVDataSource(**kwargs)
        elif source_type == "bloomberg":
            return BloombergAPIDataSource(**kwargs)
        elif source_type == "bloomberg_excel":
            return BloombergExcelDataSource(**kwargs)
        elif source_type == "bloomberg_hybrid":
            return HybridBloombergDataSource(**kwargs)
        elif source_type == "mixed_portfolio":
            return MixedPortfolioDataSource(**kwargs)
        elif source_type == "crypto":
            from _crypto_data_sources import CryptoExchangeDataSource
            return CryptoExchangeDataSource(**kwargs)
        elif source_type == "crypto_ws":
            from _crypto_data_sources import CryptoWebSocketDataSource
            return CryptoWebSocketDataSource(**kwargs)
        elif source_type == "crypto_agg":
            from _crypto_data_sources import CryptoAggregatorDataSource
            return CryptoAggregatorDataSource(**kwargs)
        elif source_type == "blockchain":
            from data_sources.blockchain_provider import BlockchainDataSource
            return BlockchainDataSource(**kwargs)
        else:
            raise ValueError(
                f"Unknown data source type: {source_type}. "
                f"Supported types: 'excel', 'csv', 'bloomberg', 'bloomberg_excel', "
                f"'bloomberg_hybrid', 'mixed_portfolio', 'crypto', 'crypto_ws', 'crypto_agg', 'blockchain'"
            )


################################################################################
# Configuration Helper
################################################################################


def load_data_from_config(config: Dict) -> pd.DataFrame:
    """
    Load data based on a configuration dictionary.

    This function provides a convenient way to load data from different
    sources based on configuration, making it easy to switch between
    data sources without changing code.

    Args:
        config: Dictionary with keys:
            - "source_type": "excel", "csv", or "bloomberg"
            - Additional keys depend on source type

    Returns:
        pd.DataFrame: Loaded data

    Example:
        >>> config = {
        ...     "source_type": "excel",
        ...     "file_path": "data/Economic_Data_2020_08_01.xlsx"
        ... }
        >>> df = load_data_from_config(config)
        >>>
        >>> # Or from environment/config file
        >>> import os
        >>> config = {
        ...     "source_type": os.getenv("DATA_SOURCE_TYPE", "excel"),
        ...     "file_path": os.getenv("DATA_FILE_PATH", "data/default.xlsx")
        ... }
        >>> df = load_data_from_config(config)
    """
    source_type = config.pop("source_type")
    source = DataSourceFactory.create(source_type, **config)
    return source.load_data()
